#!/usr/bin/env python
# -*- coding: utf-8 -*- 

import time, struct

import numpy as np

from collections import deque, namedtuple 

from openpyxl import load_workbook

num_class = 31
class_struct = namedtuple("class_t", "Class Pmin Pmax Imin Imax KI COSmin COSmax")
model_struct = namedtuple("model_t", "Class P I KI COSfi")

def init_class(class_table):
	
	class_list = [class_struct(i,float(class_table.cell(row=(i+1), column=3).value),\
				float(class_table.cell(row=(i+1), column=4).value),\
				float(class_table.cell(row=(i+1), column=5).value),\
				float(class_table.cell(row=(i+1), column=6).value),\
				float(class_table.cell(row=(i+1), column=8).value), \
				float(class_table.cell(row=(i+1), column=3).value)*1000./(440.*float(class_table.cell(row=(i+1), column=5).value)), 
				float(class_table.cell(row=(i+1), column=4).value)*1000./(440.*float(class_table.cell(row=(i+1), column=6).value)))
				for i in range(1, num_class+1) if class_table.cell(row=(i+1), column=5).value and \
													class_table.cell(row=(i+1), column=6).value and \
													class_table.cell(row=(i+1), column=8).value]
	
	#приводим в порядок интервалы cos(fi)
	for indx, cl in enumerate(class_list):
		if cl.COSmin>cl.COSmax:
			class_list[indx]=class_list[indx]._replace(COSmin=cl.COSmax)
			class_list[indx]=class_list[indx]._replace(COSmax=cl.COSmin)
			

	#for i in range(1, num_class+1):
	for cl in class_list:
		print cl.Class, cl.Pmin, cl.Pmax, cl.Imin, cl.Imax, cl.COSmin, cl.COSmax
	return class_list

def gen_rand(inta,intb):
	return np.random.uniform(inta,intb)

def init_model(class_sel,class_list):

	for cl in class_list:
		if class_sel==cl.Class:
			model = model_struct(class_sel,gen_rand(cl.Pmin,cl.Pmax),\
									gen_rand(cl.Imin,cl.Imax),\
									cl.KI,gen_rand(cl.COSmin,cl.COSmax),)
			return model
	return -1

def sort_intervals(intervals):
	arg_sort = intervals[0].argsort()
	intervals[0] = intervals[0][arg_sort]
	intervals[1] = intervals[1][arg_sort]
	return intervals

def uniqu_interval(class_list):
	intervals1 = [cl.Pmin for cl in class_list],[cl.Pmax for cl in class_list] #генерим списки интервалов
	intervals2 = [cl.COSmin for cl in class_list],[cl.COSmax for cl in class_list]
	intervals3 = [cl.Imin for cl in class_list],[cl.Imax for cl in class_list]
	KI = [cl.KI for cl in class_list] # коэффициенты тока
	
	#print np.unique(KI)
	inter_array1 = np.array(intervals1) #генерим массив интервалов
	inter_cos = np.array(intervals2)
	inter_I = np.array(intervals3)
	#print zip(inter_array[0],inter_array[1])
	#print inter_array[1]

	interval_sort = sort_intervals(inter_array1) #сортируем интервалы по нижней границе
	sort_cos = sort_intervals(inter_cos)
	sort_I = sort_intervals(inter_I)
	#print zip(interval_sort[0],interval_sort[1])
	#print interval_sort[1]
	
	cross_interval=np.argwhere(np.triu(interval_sort[1][:,None]>=interval_sort[0],1)) #вычисляем пересекающиеся интервалы
	cross_cos=np.argwhere(np.triu(sort_cos[1][:,None]>=sort_cos[0],1))
	cross_I=np.argwhere(np.triu(sort_I[1][:,None]>=sort_I[0],1))

	uniq_inter_P = (len(class_list) - len(np.unique(np.take(cross_interval,[0],1))))
	uniq_inter_COS = (len(class_list) - len(np.unique(np.take(cross_cos,[0],1))))
	uniq_inter_I = (len(class_list) - len(np.unique(np.take(cross_I,[0],1))))
	#print cross_interval
	#print len(cross_interval), len(class_list), len(np.unique(np.take(cross_interval,[0],1)))
	return uniq_inter_P + uniq_inter_COS + len(np.unique(KI)) + uniq_inter_I #число уникальных не пересекающихся значений
	

def cnt_matches_param(cl,model,u_inter):
	t_log=0.
	if model.P >=cl.Pmin and model.P <= cl.Pmax:
		t_log+=np.log(2./u_inter)
	else:
		t_log+=	np.log(1./u_inter)

	if model.KI==cl.KI:
		t_log+=np.log(2./u_inter)
	else:
		t_log+=	np.log(1./u_inter)

	if model.COSfi >= cl.COSmin and model.COSfi <= cl.COSmax:
		t_log+=np.log(2./u_inter)
	else:
		t_log+=	np.log(1./u_inter)

	if model.I >=cl.Imin and model.I <= cl.Imax:
		t_log+=np.log(2./u_inter)
	else:
		t_log+=	np.log(1./u_inter)	

	return t_log	



def classifier(model,class_list):
	print uniqu_interval(class_list) + (len(model)-2)
	u_inter = uniqu_interval(class_list) + (len(model)-2)
	prob_list=[] #список вычисляемых оценок
	prob_norm=[] #список нормированных апостериорных вероятностей 
	prob_summ=0. #сумма нормированных вероятностей по классам
	pmax=[] #максимальные вероятности
	#print len(model) 
	#вычисление апостериорных вероятностей каждого класса
	for indx,cl in enumerate(class_list):
		prob_list.append(np.log(1./len(class_list))+cnt_matches_param(cl,model,float(u_inter)))
		prob_summ+=np.exp(float(prob_list[indx]))

	#вычисляем нормированные вероятности	
	for indx,cl in enumerate(class_list):
		prob_norm.append(np.exp(prob_list[indx])/prob_summ)

		#print"class: ",cl.Class," prob: ",prob_norm[indx]

	#print type(prob_norm), len(prob_norm), np.max(prob_norm)	

	for indx, p in 	enumerate(prob_norm):
		if np.max(prob_norm)==p:
			#print p
			pmax.append((p,class_list[indx].Class))

			print"class: ",class_list[indx].Class," P: ",p
	return pmax		



	print "-"*40


def alg_ratings(Class,pmax,ar):
	if len(pmax)==1:
		if pmax[0][1]==Class:
			ar +=1

	return ar	 



if __name__=='__main__':

	
	wb = load_workbook('./class table.xlsx')
	class_table=wb.active
	class_list=init_class(class_table)
	ar=0.
	for i in class_list:

		model = init_model(i.Class,class_list)
	
		if(model!=-1):
			print "model: ",model
			pmax=classifier(model,class_list)
			ar=alg_ratings(i.Class,pmax,ar)
		else:
			print "not correct class"
		print"="*40	

	print "Rating: ",ar/len(class_list)	



	#print type(float(class_table.cell(row=2, column=3).value))
	#for i in range(1, num_class+1):
	#	print i, float(class_table.cell(row=(i+1), column=3).value)